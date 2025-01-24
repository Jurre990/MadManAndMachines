from openai import OpenAI
import openpyxl

path = "../Data/all_text_ads.xlsx"

# best working prompt
prompts = ["Assign one or more codes to the advertisements following the rules in the codebook. Please give each one as a list and only mention the codes that apply. Seperate the result of each as with =====."]

# the codebook explains the definitions of all codes
codebook = ""
with open("../Data/codebook.txt", encoding='utf-8') as file:
    codebook = file.read()


#loop through the ads
i = 199
while i<1129:
    print("----------- "+str(i)+" -------------")

    # load 5 different ads
    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active
    text1 = sheet_obj.cell(row=i, column=2).value
    text2 = sheet_obj.cell(row=i+1, column=2).value
    text3 = sheet_obj.cell(row=i+2, column=2).value
    text4 = sheet_obj.cell(row=i+3, column=2).value
    text5 = sheet_obj.cell(row=i+4, column=2).value

    # connect with Chat-GPT
    client = OpenAI()

    # send the codebook, 5 ads and the prompt to get a response
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text",
                     "text": "Codebook: " + codebook},
                    {
                        "type": "text",
                        "text": "Advertisement: " + text1,
                    },
                    {
                        "type": "text",
                        "text": "Advertisement: " + text2,
                    },
                    {
                        "type": "text",
                        "text": "Advertisement: " + text3,
                    },
                    {
                        "type": "text",
                        "text": "Advertisement: " + text4,
                    },
                    {
                        "type": "text",
                        "text": "Advertisement: " + text5,
                    },
                    {
                        "type": "text",
                        "text": prompts[0],
                    }
                ],
            },
        ],
    )

    print(response)

    responses = response.choices[0].message.content.strip().split("=====")
    print(len(responses))

    # check if there are actually 5 answers given, because that is often not the case
    if len(responses) != 5:
        #try the ad again
        continue

    # note the mentioned codes in the response
    for text_response in responses:
        codes = ["E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3", "I-4", "I-5", "I-6", "Ro-1",
                 "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

        code_list = []
        for code in codes:
            if code in text_response:
                code_list.append(code)
        if code_list == []:
            code_list = "N"

    print(code_list)

    i += 5
