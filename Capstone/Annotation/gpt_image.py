from openai import OpenAI
import openpyxl

path = "../Data/image_ads.xlsx"

# 3 different prompts that were tried:
prompts = ["Assign one or more codes to the advertisement following the rules in the codebook. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences.",
           "You are an employee of meta and have been tasked with checking a number of political ads for possible violations that could make the platform unsafe. Only provide a code if you are completely sure. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences.",
           "You are a PHD political science student researching the amount of dangerous political ads on social media. Give a code according to the codebook to the ad if it falls under the code with significant probability. Most ads will not get a code, then give an N. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences."]

# the codebook explains the definitions of all codes
codebook = ""
with open("../Data/codebook.txt", encoding='utf-8') as file:
    codebook = file.read()


# load the text and image links from the ads from the excel sheet into a list
wb = openpyxl.load_workbook(path)
sheet_obj = wb.active
text = []
links = []
for i in range(1000):
    text.append(sheet_obj.cell(row=i + 4, column=2).value)
    links.append(sheet_obj.cell(row=i + 4, column=3).value)
    if text[-1] == None:
        text[-1] = ""

#loop through the ads
i = 0
while i<200:
    print("----------- "+str(i)+" -------------")

    print(links[i])

    # connect with Chat-GPT
    client = OpenAI()

    # send the codebook, ad and prompt to get a response
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text",
                     "text": "Codebook: " + codebook + " Advertisement: " + text[i] },
                    {
                        "type": "image_url",
                        "image_url": {"url": links[i]},
                    },
                    {
                        "type": "text",
                        "text": prompts[0],
                    }
                ],
            },
        ],
    )

    print(response)

    # note the mentioned codes in the response
    codes = ["E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3", "I-4", "I-5", "I-6", "Ro-1:",
             "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

    code_list = []
    text_response = response.choices[0].message.content.strip()
    for code in codes:
        if code in text_response:
            print("Code: " + code)
            code_list.append(code)
    if code_list == []:
        # check if the response give a code or not
        if "N" in text_response or "violations" in text_response or "code" in text_response:
            code_list = "N"
        else:
            # try the same ad again
            code_list = "error"
            continue

    # write the codes to the excell sheet
    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active
    sheet_obj.cell(row=i + 4, column=4).value = ','.join(code_list)
    sheet_obj.cell(row=i + 4, column=5).value = response.choices[0].message.content.strip()

    #wb.save(path)
    # ^ uncomment to save new data

    i += 1
