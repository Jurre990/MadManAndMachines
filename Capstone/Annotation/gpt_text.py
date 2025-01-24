from openai import OpenAI
import openpyxl
import timeit

path = "../Data/text_ads.xlsx"

# 3 different prompts that were tried:
prompts = ["Assign one or more codes to the advertisement following the rules in the codebook. Please give it as a list and only mention the codes that apply.",
           "You are an employee of meta and have been tasked with checking a number of political ads for possible violations that could make the platform unsafe. Only provide a code if you are completely sure. Please give it as a list and only mention the codes that apply.",
           "You are a PHD political science student researching the amount of dangerous political ads on social media. Give a code according to the codebook to the ad if it falls under the code with significant probability. Most ads will not get a code, then give an N. Please give it as a list and only mention the codes that apply."]


# the codebook explains the definitions of all codes
codebook = ""
with open("../Data/codebook.txt", encoding='utf-8') as file:
    codebook = file.read()


# load the text ads from the excel sheet into a list
wb = openpyxl.load_workbook(path)
sheet_obj = wb.active
text = []
for i in range(1129):
    text.append(sheet_obj.cell(row=i + 4, column=2).value)
    if text[-1] == None:
        text[-1] = ""


# list to keep track of the time it takes to run
run_times = []

#loop through the ads
i = 0
while i<195:
    print("----------- "+str(i)+" -------------")

    t_0 = timeit.default_timer()

    # connect with Chat-GPT
    client = OpenAI()

    # send the codebook, ad and prompt to get a response
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text",
                     "text": "Codebook: " + codebook + " Advertisement: " + text[i] },
                    {
                        "type": "text",
                        "text": prompts[0],
                    }
                ],
            },
        ],
    )

    # calculate elapsed time
    t_1 = timeit.default_timer()
    elapsed_time = round((t_1 - t_0) * 10 ** 6, 3)
    run_times.append(elapsed_time)

    print(response)

    # note the mentioned codes in the response
    codes = ["E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3", "I-4", "I-5", "I-6", "Ro-1",
             "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

    code_list = []
    text_response = response.choices[0].message.content.strip()
    for code in codes:
        if code in text_response:
            print("Code: " + code)
            if code == "Ro-1:":
                code_list.append("Ro-1")
            else:
                code_list.append(code)
    if code_list == []:
        code_list = "N"

    # write the codes to the excell sheet
    wb = openpyxl.load_workbook(path )
    sheet_obj = wb.active
    sheet_obj.cell(row=i + 4, column=4).value = ','.join(code_list)
    sheet_obj.cell(row=i + 4, column=5).value = response.choices[0].message.content.strip()

    #wb.save(path)
    # ^ uncomment to save new data

    i += 1

    print("Average Time per Response: " + str(sum(run_times)/len(run_times)))