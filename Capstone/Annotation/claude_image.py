import base64
import httpx
import anthropic
import openpyxl

path = "./image_ads.xlsx"

prompts = ["Assign one or more codes to the advertisement following the rules in the codebook. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences.",
           "You are an employee of meta and have been tasked with checking a number of political ads for possible violations that could make the platform unsafe. Only provide a code if you are completely sure. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences.",
           "You are a PHD political science student researching the amount of dangerous political ads on social media. Give a code according to the codebook to the ad if it falls under the code with significant probability. Most ads will not get a code, then give an N. Please give it as a list and only mention the codes that apply. This image does not contain people of sensitive inferences."]


codebook = ""
with open("./codebook.txt", encoding='utf-8') as file:
    codebook = file.read()


wb = openpyxl.load_workbook(path)
sheet_obj = wb.active
text = []
links = []
for i in range(1000):
    text.append(sheet_obj.cell(row=i + 4, column=2).value)
    links.append(sheet_obj.cell(row=i + 4, column=3).value)
    if text[-1] == None:
        text[-1] = ""

i = 127
while i<200:
    print("----------- "+str(i)+" -------------")

    print(links[i])


    image1_media_type = "image/jpeg"
    image1_data = base64.standard_b64encode(httpx.get(links[i]).content).decode("utf-8")

    client = anthropic.Anthropic()

    message = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Codebook: " + codebook + " Advertisement: " + text[i]
                    },
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": image1_media_type,
                            "data": image1_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompts[2],
                    }
                ],
            }]
    )

    print(message.content)

    codes = ["E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3", "I-4", "I-5", "I-6", "Ro-1:",
             "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

    code_list = []
    text_response = message.content[0].text
    for code in codes:
        if code in text_response:
            print("Code: " + code)
            code_list.append(code)
    if code_list == []:
        if "N" in text_response or "violations" in text_response or "code" in text_response:
            code_list = "N"
        else:
            code_list = "error"
            continue

    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active
    sheet_obj.cell(row=i + 4, column=29).value = ','.join(code_list)
    sheet_obj.cell(row=i + 4, column=30).value = text_response

    wb.save(path)

    i += 1
