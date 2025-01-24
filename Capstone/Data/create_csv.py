import openpyxl
import pandas as pd

path = "./all_text_ads.xlsx"

codes_index = ["N", "E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3",
               "I-4", "I-5", "I-6", "Ro-1",
               "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

dict = { "text": [], "codes": []}

# open the excel sheet and put the data in a dict
wb = openpyxl.load_workbook(path)
sheet_obj = wb.active
for i in range(4,1126,1):
    text = sheet_obj.cell(row=i, column=2).value
    codes = [codes_index.index(c) for c in sheet_obj.cell(row=i, column=4).value.split(",") ]
    dict["text"].append(text)
    dict["codes"].append(codes)

# convert the dict to a pandas dataframe
df = pd.DataFrame(dict)
print(df)

# store it as a .csv
df.to_csv('text_data.csv')
