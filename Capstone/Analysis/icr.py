import openpyxl
import numpy as np
import krippendorff
import matplotlib.pyplot as plt

# a list of all possible codes where the index will be used as code ID
codes_index = ["N", "E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3",
               "I-4", "I-5", "I-6", "Ro-1",
               "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]

def inter_coder_reliability(path, column1, column2):

    #============= read data =============
    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active
    coder_1 = [] # store all the codes from the first coder
    coder_2 = [] # store all the codes from the second coder
    for i in range(200):
        if sheet_obj.cell(row=i + 4, column=column1).value != None and sheet_obj.cell(row=i + 4,
                                                                                      column=column2).value != None:
            coder_1.append(sheet_obj.cell(row=i + 4, column=column1).value)
            coder_2.append(sheet_obj.cell(row=i + 4, column=column2).value)

    #============= create value_counts ==================


    # create a list with a number for all possible codes to keep track of how often
    # that code is mentioned by the coders
    value_counts = []
    for i in range(len(coder_1)):
        value_counts.append([0 for n in range(len(codes_index))])

    # use the list of all codes to get an index and increase the value of the number
    # at that index in value_counts
    for i in range(len(coder_1)):
        coder_1[i] = coder_1[i].split(",")
        for n in range(len(coder_1[i])):
            value_counts[i][codes_index.index(coder_1[i][n].replace(u'\xa0', u'').replace(' ',''))] += 1

    for i in range(len(coder_2)):
        coder_2[i] = coder_2[i].split(",")
        for n in range(len(coder_2[i])):
            value_counts[i][codes_index.index(coder_2[i][n].replace(u'\xa0', u'').replace(' ',''))] += 1

    #=============== split value_count into more rows ================

    # make sure that on every row the sum is two by splitting rows into multiple
    new_value_counts = []
    for i in range(len(value_counts)):
        row_counter = 0
        for n in range(len(codes_index)):

            # keep two of the same code on the same row
            if value_counts[i][n] == 2:
                new_value_counts.append([0 for m in range(len(codes_index))])
                new_value_counts[-1][n] = 2
            elif value_counts[i][n] == 1:
                if row_counter == 0:
                    # create a new row
                    new_value_counts.append([0 for m in range(len(codes_index))])
                    new_value_counts[-1][n] = 1
                    row_counter = 1
                else:
                    # use the last created new row
                    new_value_counts[-1][n] = 1
                    row_counter = 0
        if row_counter == 1:
            # add an extra N to the last row if the sum is only 1
            new_value_counts[-1][0] = 1


    #================ create binary value counts ================

    # for the binary value counts, only keep track of code/no code
    binary_value_counts = []
    for i in range(len(coder_1)):
        temp = [0,0]
        if coder_1[i]!=["N"]:
            temp[1] += 1
        else:
            temp[0] += 1
        if coder_2[i]!=["N"]:
            temp[1] += 1
        else:
            temp[0] += 1
        binary_value_counts.append(temp)

    #================ calculate krippendorff ==================

    value_counts = np.array(new_value_counts)

    alpha = krippendorff.alpha(value_counts=value_counts,
                               level_of_measurement="nominal")

    alpha_binary = krippendorff.alpha(value_counts=binary_value_counts,
                               level_of_measurement="nominal")

    return [alpha, alpha_binary]


text_gpt_1 = inter_coder_reliability("../Data/text_ads.xlsx", 4, 24)
text_claude_1 = inter_coder_reliability("../Data/text_ads.xlsx", 7, 24)
text_gpt_2 = inter_coder_reliability("../Data/text_ads.xlsx", 10, 24)
text_claude_2 = inter_coder_reliability("../Data/text_ads.xlsx", 13, 24)
text_gpt_3 = inter_coder_reliability("../Data/text_ads.xlsx", 16, 24)
text_claude_3 = inter_coder_reliability("../Data/text_ads.xlsx", 19, 24)
text_gpt_1_self = inter_coder_reliability("../Data/text_ads.xlsx", 4, 26)
text_claude_1_self = inter_coder_reliability("../Data/text_ads.xlsx", 7, 29)
text_human = inter_coder_reliability("../Data/text_ads.xlsx", 22, 23)

image_gpt_1 = inter_coder_reliability("../Data/image_ads.xlsx", 4, 24)
image_claude_1 = inter_coder_reliability("../Data/image_ads.xlsx", 7, 24)
image_gpt_2 = inter_coder_reliability("../Data/image_ads.xlsx", 10, 24)
image_claude_2 = inter_coder_reliability("../Data/image_ads.xlsx", 13, 24)
image_gpt_3 = inter_coder_reliability("../Data/image_ads.xlsx", 16, 24)
image_claude_3 = inter_coder_reliability("../Data/image_ads.xlsx", 19, 24)
image_human = inter_coder_reliability("../Data/image_ads.xlsx", 22, 23)

video_gpt_1 = inter_coder_reliability("../Data/video_ads.xlsx", 4, 24)
video_claude_1 = inter_coder_reliability("../Data/video_ads.xlsx", 7, 24)
video_gpt_2 = inter_coder_reliability("../Data/video_ads.xlsx", 10, 24)
video_claude_2 = inter_coder_reliability("../Data/video_ads.xlsx", 13, 24)
video_gpt_3 = inter_coder_reliability("../Data/video_ads.xlsx", 16, 24)
video_claude_3 = inter_coder_reliability("../Data/video_ads.xlsx", 19, 24)
video_human = inter_coder_reliability("../Data/video_ads.xlsx", 22, 23)

print("--------- Text ---------")
print("GPT Prompt 1: " + str(text_gpt_1))
print("GPT Prompt 2: " + str(text_gpt_2))
print("GPT Prompt 3: " + str(text_gpt_3))
print("Claude Prompt 1: " + str(text_claude_1))
print("Claude Prompt 2: " + str(text_claude_2))
print("Claude Prompt 3: " + str(text_claude_3))
print("GPT Prompt 1 Self: " + str(text_gpt_1_self))
print("Claude Prompt 1 Self: " + str(text_claude_1_self))

print("--------- Image ---------")
print("GPT Prompt 1: " + str(image_gpt_1))
print("GPT Prompt 2: " + str(image_gpt_2))
print("GPT Prompt 3: " + str(image_gpt_3))
print("Claude Prompt 1: " + str(image_claude_1))
print("Claude Prompt 2: " + str(image_claude_2))
print("Claude Prompt 3: " + str(image_claude_3))

print("--------- Video ---------")
print("GPT Prompt 1: " + str(video_gpt_1))
print("GPT Prompt 2: " + str(video_gpt_2))
print("GPT Prompt 3: " + str(video_gpt_3))
print("Claude Prompt 1: " + str(video_claude_1))
print("Claude Prompt 2: " + str(video_claude_2))
print("Claude Prompt 3: " + str(video_claude_3))


# ============ create a bar chart for the krippendorff scores ===========
barWidth = 0.25

prompt1 = [text_gpt_1[0], text_claude_1[0], image_gpt_1[0], image_claude_1[0], video_gpt_1[0], video_claude_1[0]]
prompt2 = [text_gpt_2[0], text_claude_2[0], image_gpt_2[0], image_claude_2[0], video_gpt_2[0], video_claude_2[0]]
prompt3 = [text_gpt_3[0], text_claude_3[0], image_gpt_3[0], image_claude_3[0], video_gpt_3[0], video_claude_3[0]]

br1 = np.arange(6)
br2 = [x + barWidth for x in br1]
br3 = [x + barWidth for x in br2]

plt.bar(br1, prompt1, color ='crimson', width = barWidth,
        edgecolor ='grey', label ='Prompt 1')
plt.bar(br2, prompt2, color ='gold', width = barWidth,
        edgecolor ='grey', label ='Prompt 2')
plt.bar(br3, prompt3, color ='darkcyan', width = barWidth,
        edgecolor ='grey', label ='Prompt 3')

plt.xlabel('Model and ad type', fontweight ='bold', fontsize = 10)
plt.ylabel('Krippendorff Alpha Coefficient', fontweight ='bold', fontsize = 10)
plt.xticks([r + barWidth for r in range(6)],
        ['Chat-GPT\nText', 'Claude\nText', 'Chat-GPT\nImage', 'Claude\nImage',
               'Chat-GPT\nVideo', 'Claude\nVideo'])

plt.legend()
plt.title("Inter Coder Reliability", fontweight ='bold')
plt.tight_layout()
plt.show()


# ============ create a bar chart for the binary krippendorff scores ===========

prompt1_b = [text_gpt_1[1], text_claude_1[1], image_gpt_1[1], image_claude_1[1], video_gpt_1[1], video_claude_1[1]]
prompt2_b = [text_gpt_2[1], text_claude_2[1], image_gpt_2[1], image_claude_2[1], video_gpt_2[1], video_claude_2[1]]
prompt3_b = [text_gpt_3[1], text_claude_3[1], image_gpt_3[1], image_claude_3[1], video_gpt_3[1], video_claude_3[1]]

plt.bar(br1, prompt1_b, color ='crimson', width = barWidth,
        edgecolor ='grey', label ='Prompt 1')
plt.bar(br2, prompt2_b, color ='gold', width = barWidth,
        edgecolor ='grey', label ='Prompt 2')
plt.bar(br3, prompt3_b, color ='darkcyan', width = barWidth,
        edgecolor ='grey', label ='Prompt 3')

plt.xlabel('Model and ad type', fontweight ='bold', fontsize = 10)
plt.ylabel('Krippendorff Alpha Coefficient', fontweight ='bold', fontsize = 10)
plt.xticks([r + barWidth for r in range(6)],
        ['Chat-GPT\nText', 'Claude\nText', 'Chat-GPT\nImage', 'Claude\nImage',
               'Chat-GPT\nVideo', 'Claude\nVideo'])

plt.legend()
plt.title("Binary Inter Coder Reliability", fontweight ='bold')
plt.tight_layout()
plt.show()

# ============ create a bar chart for the reliability of coders ===========

barlist = plt.bar(["Human\nText", "Human\nImage", "Human\nVideo", "Chat-GPT\nText", "Claude\nText"], [text_human[0], image_human[0], video_human[0], text_gpt_1_self[0], text_claude_1_self[0]], color="crimson")
barlist[3].set_color('darkcyan')
barlist[4].set_color('darkcyan')
plt.title("Coder Reliability", fontweight ='bold')
plt.xlabel('Human/model and ad type', fontweight ='bold', fontsize = 10)
plt.ylabel('Krippendorff Alpha Coefficient of 2 iterations', fontweight ='bold', fontsize = 10)
plt.tight_layout()
plt.show()