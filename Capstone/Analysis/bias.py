import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# keep track of what party the ads are in favor of
alignment = []

# R: Republican, D: Democratic, N: Neutral
parties = ["R", "D", "N"]

# keep track of how many ads are annotated by the model, split by alignment
gpt1 = [0, 0, 0]
gpt2 = [0, 0, 0]
gpt3 = [0, 0, 0]
claude1 = [0, 0, 0]
claude2 = [0, 0, 0]
claude3 = [0, 0, 0]
human = [0, 0, 0]

# load the image annotations from the excel sheet
wb = openpyxl.load_workbook("../Data/image_ads.xlsx")
sheet_obj = wb.active
for i in range(4, 199, 1):
    alignment.append(sheet_obj.cell(row=i, column=26).value.replace(u'\xa0', u'').replace(' ',''))
    if sheet_obj.cell(row=i, column=4).value != "N":
        gpt1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=7).value != "N":
        claude1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=10).value != "N":
        gpt2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=13).value != "N":
        claude2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=16).value != "N":
        gpt3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=19).value != "N":
        claude3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=24).value != "N":
        human[parties.index(alignment[-1])] += 1

# load the video annotations from the excel sheet
wb = openpyxl.load_workbook("../Data/video_ads.xlsx")
sheet_obj = wb.active
for i in range(4, 173, 1):
    alignment.append(sheet_obj.cell(row=i, column=26).value.replace(u'\xa0', u'').replace(' ',''))
    if sheet_obj.cell(row=i, column=4).value != "N":
        gpt1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=7).value != "N":
        claude1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=10).value != "N":
        gpt2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=13).value != "N":
        claude2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=16).value != "N":
        gpt3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=19).value != "N":
        claude3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=24).value != "N":
        human[parties.index(alignment[-1])] += 1

# load the text annotations from the excel sheet
wb = openpyxl.load_workbook("../Data/text_ads.xlsx")
sheet_obj = wb.active
for i in range(4, 173, 1):
    alignment.append(sheet_obj.cell(row=i, column=32).value.replace(u'\xa0', u'').replace(' ',''))
    if sheet_obj.cell(row=i, column=4).value != "N":
        gpt1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=7).value != "N":
        claude1[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=10).value != "N":
        gpt2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=13).value != "N":
        claude2[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=16).value != "N":
        gpt3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=19).value != "N":
        claude3[parties.index(alignment[-1])] += 1
    if sheet_obj.cell(row=i, column=24).value != "N":
        human[parties.index(alignment[-1])] += 1

# count how many ads there in each category
y = [alignment.count("R"), alignment.count("D"), alignment.count("N")]

# calculate the ratio of the total per category
ratio = [y[0] / len(alignment), y[1] / len(alignment), y[2] / len(alignment)]

# plot the amount per category in a pie chart
plt.title("Ads Political Ideology", fontweight ='bold')
plt.pie(y, labels=["Republican " + str(round(100 * ratio[0], 1)) + "%",
                   "Democratic " + str(round(100 * ratio[1], 1)) + "%",
                   "Neutral " + str(round(100 * ratio[2], 1)) + "%"],
        colors=["orangered", "dodgerblue", "slategray"])
plt.show()

# ================ bar chart of the amount of annotations per category per model =============

barWidth = 0.25

r_ann = [gpt1[0], claude1[0], gpt2[0], claude2[0], gpt3[0], claude3[0], human[0]]
d_ann = [gpt1[1], claude1[1], gpt2[1], claude2[1], gpt3[1], claude3[1], human[1]]
n_ann = [gpt1[2], claude1[2], gpt2[2], claude2[2], gpt3[2], claude3[2], human[2]]

br1 = np.arange(7)
br2 = [x + barWidth for x in br1]
br3 = [x + barWidth for x in br2]

plt.bar(br1, r_ann, color ='orangered', width = barWidth,
        edgecolor ='grey', label ='Republican')
plt.bar(br2, d_ann, color ='dodgerblue', width = barWidth,
        edgecolor ='grey', label ='Democratic')
plt.bar(br3, n_ann, color ='slategray', width = barWidth,
        edgecolor ='grey', label ='Neutral')

plt.xlabel('Model and prompt', fontweight ='bold', fontsize = 10)
plt.ylabel('Amount of ads classified as violation', fontweight ='bold', fontsize = 10)
plt.xticks([r + barWidth for r in range(7)],
        ['Chat-GPT\nPrompt 1', 'Claude\nPrompt 1', 'Chat-GPT\nPrompt 2', 'Claude\nPrompt 2',
               'Chat-GPT\nPrompt 3', 'Claude\nPrompt 3', 'Human'])

plt.legend()
plt.title("Annotated ads per ideology", fontweight ='bold')
plt.tight_layout()
plt.show()


# ================ bar chart of the normalized amount of annotations per category per model =============

# normalize by the amount each model has annotated
gpt1_norm = [float(i)/sum(gpt1) for i in gpt1]
gpt2_norm = [float(i)/sum(gpt2) for i in gpt2]
gpt3_norm = [float(i)/sum(gpt3) for i in gpt3]
claude1_norm = [float(i)/sum(claude1) for i in claude1]
claude2_norm = [float(i)/sum(claude2) for i in claude2]
claude3_norm = [float(i)/sum(claude3) for i in claude3]
human_norm = [float(i)/sum(human) for i in human]

# normalize by how many ads of that ideology there are
r_ann_norm = [gpt1_norm[0]/ratio[0], claude1_norm[0]/ratio[0], gpt2_norm[0]/ratio[0], claude2_norm[0]/ratio[0], gpt3_norm[0]/ratio[0], claude3_norm[0]/ratio[0], human_norm[0]/ratio[0]]
d_ann_norm = [gpt1_norm[1]/ratio[1], claude1_norm[1]/ratio[1], gpt2_norm[1]/ratio[1], claude2_norm[1]/ratio[1], gpt3_norm[1]/ratio[1], claude3_norm[1]/ratio[1], human_norm[1]/ratio[1]]
n_ann_norm = [gpt1_norm[2]/ratio[2], claude1_norm[2]/ratio[2], gpt2_norm[2]/ratio[2], claude2_norm[2]/ratio[2], gpt3_norm[2]/ratio[2], claude3_norm[2]/ratio[2], human_norm[2]/ratio[2]]

plt.bar(br1, r_ann_norm, color ='orangered', width = barWidth,
        edgecolor ='grey', label ='Republican')
plt.bar(br2, d_ann_norm, color ='dodgerblue', width = barWidth,
        edgecolor ='grey', label ='Democratic')
plt.bar(br3, n_ann_norm, color ='slategray', width = barWidth,
        edgecolor ='grey', label ='Neutral')

plt.xlabel('Model and prompt', fontweight ='bold', fontsize = 10)
plt.ylabel('Amount of ads classified as violation, normalized', fontweight ='bold', fontsize = 10)
plt.xticks([r + barWidth for r in range(7)],
        ['Chat-GPT\nPrompt 1', 'Claude\nPrompt 1', 'Chat-GPT\nPrompt 2', 'Claude\nPrompt 2',
               'Chat-GPT\nPrompt 3', 'Claude\nPrompt 3', 'Human'])

plt.legend(loc='lower left')
plt.title("Annotation Bias", fontweight ='bold')
plt.tight_layout()
plt.show()


# ======================= plot pie chart of how many times each code is used in the annotations ==================

codes_index = ["N", "E-2", "E-3", "E-4", "E-5", "E-6", "E-7", "E-8", "C-1", "C-2", "C-3", "C-4", "I-1", "I-2", "I-3",
               "I-4", "I-5", "I-6", "Ro-1",
               "Ro-7", "Ro-9", "Ro-10", "D-1", "D-2", "D-3", "D-4", "D-5", "D-6", "D-7", "D-8", "D-9", "D-10"]


def count_codes(path, column): # count the amount of times each code is used
    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active

    coder = []
    for i in range(200):
        if sheet_obj.cell(row=i + 4, column=column).value != None:
            coder.append(sheet_obj.cell(row=i + 4, column=column).value)

    # codes_count is a list with a number for every type of code to count how often it is used
    codes_count = [0 for _ in range(len(codes_index))]
    for i in range(len(coder)):
        coder[i] = coder[i].split(",")
        for n in range(len(coder[i])):
            codes_count[codes_index.index(coder[i][n].replace(u'\xa0', u'').replace(' ',''))] += 1

    # remove the N
    codes_count.pop(0)
    return codes_count

text_gpt_1 = count_codes("../Data/text_ads.xlsx", 4)
text_claude_1 = count_codes("../Data/text_ads.xlsx", 7)
text_gpt_2 = count_codes("../Data/text_ads.xlsx", 10)
text_claude_2 = count_codes("../Data/text_ads.xlsx", 13)
text_gpt_3 = count_codes("../Data/text_ads.xlsx", 16)
text_claude_3 = count_codes("../Data/text_ads.xlsx", 19)
text_gpt_1_self = count_codes("../Data/text_ads.xlsx", 4)
text_human = count_codes("../Data/text_ads.xlsx", 22)

image_gpt_1 = count_codes("../Data/image_ads.xlsx", 4)
image_claude_1 = count_codes("../Data/image_ads.xlsx", 7)
image_gpt_2 = count_codes("../Data/image_ads.xlsx", 10)
image_claude_2 = count_codes("../Data/image_ads.xlsx", 13)
image_gpt_3 = count_codes("../Data/image_ads.xlsx", 16)
image_claude_3 = count_codes("../Data/image_ads.xlsx", 19)
image_human = count_codes("../Data/image_ads.xlsx", 22)

video_gpt_1 = count_codes("../Data/video_ads.xlsx", 4)
video_claude_1 = count_codes("../Data/video_ads.xlsx", 7)
video_gpt_2 = count_codes("../Data/video_ads.xlsx", 10)
video_claude_2 = count_codes("../Data/video_ads.xlsx", 13)
video_gpt_3 = count_codes("../Data/video_ads.xlsx", 16)
video_claude_3 = count_codes("../Data/video_ads.xlsx", 19)
video_human = count_codes("../Data/video_ads.xlsx", 22)

codes_index.pop(0)



def create_sub_pie(codes_list, title, location):
    # function to create a pie chart of the given codes_list

    ax1 = plt.subplot2grid((3,3),location)
    codes_list = [float(i)/sum(codes_list) for i in codes_list]
    new_text = []
    new_codes_index = []
    for i in range(len(codes_list)):
        # only keep the code if it is used
        if codes_list[i] > 0:
            new_text.append(codes_list[i])
            # remove the label if the fraction is too small
            if codes_list[i] < 0.03:
                new_codes_index.append("")
            else:
                new_codes_index.append(codes_index[i])

    plt.title(title)
    plt.pie(new_text, labels=new_codes_index, colors=['tomato', 'cornflowerblue', 'gold', 'orchid', 'green'] )

fig = plt.figure(figsize= (8, 12))

create_sub_pie(text_gpt_1, "Chat-GPT\nPrompt 1", (0,0))
create_sub_pie(text_gpt_2, "Chat-GPT\nPrompt 2", (0,1))
create_sub_pie(text_gpt_3, "Chat-GPT\nPrompt 3", (0,2))
create_sub_pie(text_claude_1, "Claude\nPrompt 1", (1,0))
create_sub_pie(text_claude_2, "Claude\nPrompt 2", (1,1))
create_sub_pie(text_claude_3, "Claude\nPrompt 3", (1,2))
create_sub_pie(text_human, "Human", (2,0))

plt.suptitle("Text Ads", y=0.1, fontweight ='bold')
plt.tight_layout()
plt.show()

fig = plt.figure(figsize= (8, 12))

create_sub_pie(image_gpt_1, "Chat-GPT\nPrompt 1", (0,0))
create_sub_pie(image_gpt_2, "Chat-GPT\nPrompt 2", (0,1))
create_sub_pie(image_gpt_3, "Chat-GPT\nPrompt 3", (0,2))
create_sub_pie(image_claude_1, "Claude\nPrompt 1", (1,0))
create_sub_pie(image_claude_2, "Claude\nPrompt 2", (1,1))
create_sub_pie(image_claude_3, "Claude\nPrompt 3", (1,2))
create_sub_pie(image_human, "Human", (2,0))

plt.suptitle("Image Ads", y=0.1, fontweight ='bold')
plt.tight_layout()
plt.show()

fig = plt.figure(figsize= (8, 12))

create_sub_pie(video_gpt_1, "Chat-GPT\nPrompt 1", (0,0))
create_sub_pie(video_gpt_2, "Chat-GPT\nPrompt 2", (0,1))
create_sub_pie(video_gpt_3, "Chat-GPT\nPrompt 3", (0,2))
create_sub_pie(video_claude_1, "Claude\nPrompt 1", (1,0))
create_sub_pie(video_claude_2, "Claude\nPrompt 2", (1,1))
create_sub_pie(video_claude_3, "Claude\nPrompt 3", (1,2))
create_sub_pie(video_human, "Human", (2,0))

plt.suptitle("Video Ads", y=0.1, fontweight ='bold')
plt.tight_layout()
plt.show()