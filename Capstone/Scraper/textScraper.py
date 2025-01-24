from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl

def getAd(url):
    service = Service()
    #options = webdriver.ChromeOptions()
    #driver = webdriver.Chrome(service=service, options=options)
    options = Options()
    options.add_argument("--headless=new")

    options.add_argument("user-data-dir=C:/Users/jurre/AppData/Local/Google/Chrome/User Data")
    driver = webdriver.Chrome(options=options)

    driver.get(url)

    delay = 3  # seconds

    try:
        myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.XPATH,
                                                                                    '/html/body/div[5]/div[1]/div[1]/div/div/div/div/div[2]/div[1]/div[2]/div/div/div/div/div/div[3]/div/div/div[2]/div/span/div/div')))
        text = myElem.text

    except TimeoutException:
        source = str(driver.page_source.encode("utf-8"))
        index1 = source.find(
            '<div class="_4ik4 _4ik5" style="line-height: 16px; max-height: 112px; -webkit-line-clamp: 7;">') + 94
        index2 = source.find('</div>', index1)
        text = source[index1:index2:1]

    #print(text)
    driver.quit()
    if text.endswith('style="top: -1px;">'):
        return ""

    return text


def getAdLinks(path):
    # get the links to all the ads and put them in a list
    wb = openpyxl.load_workbook(path)
    sheet_obj = wb.active
    links = []
    for i in range(6000):
        links.append(sheet_obj.cell(row=i + 4, column=1).value)
    return links


# list with all the links to the ad pages
adLinks = getAdLinks("text_annotation.xlsx")

# loop through the links
for i in range(0, 200, 1):
    print("------------- index "+str(i)+" -------------")
    print(adLinks[i])
    # scrape the text from the webpage
    text = getAd(adLinks[i])
    print(text)

    # store the text in the excel sheet
    wb = openpyxl.load_workbook("text_annotation.xlsx")
    sheet_obj = wb.active
    sheet_obj.cell(row=i + 4, column=2).value = text
    wb.save("text_annotation.xlsx")